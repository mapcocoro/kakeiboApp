#!/usr/bin/env python3
"""
Excelの家計簿データをアプリ用のCSV形式に変換するスクリプト
"""

import openpyxl
import csv
import re
from datetime import datetime

def categorize_item(place, description):
    """
    場所と商品名からカテゴリを自動判定
    """
    text = f"{place or ''} {description or ''}".lower()

    # ローンや保険などの固定費
    if any(word in text for word in ['ローン', '保険', 'loan', 'insurance', 'メットライフ',
                                      'アクサ', '車両保険', 'aig']):
        return '固定費'

    # 通信費（固定費に含める）
    if any(word in text for word in ['ソネット', 'モバイル', '携帯', 'スマホ', 'wifi',
                                      'インターネット', 'ネット', '通信']):
        return '固定費'

    # 光熱費
    if any(word in text for word in ['電気', '水道', 'ガス', '光熱', 'でんき', 'みず',
                                      '東京電力', '東京ガス', '下水', '水道局']):
        return '光熱費'

    # ジム・娯楽
    if any(word in text for word in ['ジム', 'アッティーボ', '映画', 'レジャー', '娯楽']):
        return '娯楽費'

    # 交通費
    if any(word in text for word in ['ガソリン', '電車', 'バス', '交通', 'えき', 'タクシー',
                                      'suica', 'パスモ']):
        return '交通費'

    # 医療費
    if any(word in text for word in ['病院', '薬', '医療', 'クリニック', 'びょういん',
                                      'ドラッグ', '薬局']):
        return '医療費'

    # 食費（デフォルト）
    # スーパー、食材、飲食店など
    return '食費'

def parse_date(sheet_name, day, year_offset=0):
    """
    シート名（YY.MM形式）と日付から完全な日付を生成
    """
    try:
        # シート名から年月を抽出
        match = re.match(r'(\d+)\.(\d+)', sheet_name)
        if not match:
            return None

        year_short = int(match.group(1))
        month = int(match.group(2))

        # 20XX年に変換
        year = 2000 + year_short + year_offset

        # 日付が数値でない場合はスキップ
        if not isinstance(day, (int, float)):
            return None

        day = int(day)

        # 日付の妥当性チェック
        if day < 1 or day > 31:
            return None

        # 日付を生成
        date_str = f"{year:04d}-{month:02d}-{day:02d}"

        # 日付の妥当性を確認
        try:
            datetime.strptime(date_str, '%Y-%m-%d')
            return date_str
        except ValueError:
            return None

    except Exception as e:
        return None

def extract_expenses(file_path):
    """
    Excelファイルから支出データを抽出
    """
    print("Excelファイルを読み込んでいます...")
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

    expenses = []

    # 月次シートのみを対象（集計シートは除外）
    month_sheets = [name for name in wb.sheetnames
                   if re.match(r'\d+\.\d+', name)]

    print(f"\n{len(month_sheets)}個の月次シートを処理します...")

    for sheet_name in month_sheets:
        print(f"処理中: {sheet_name}")
        ws = wb[sheet_name]

        # データが存在する行を探す
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 8:
                continue

            # 列の位置（最初の分析結果から）
            # 列1: 月/日付, 列2: タスク, 列3: チェック, 列4: 日, 列5: 曜日
            # 列6: 場所, 列7: 価格, 列8: 商品名

            try:
                day = row[3]  # 日
                place = row[5]  # 場所
                amount = row[6]  # 価格
                description = row[7]  # 商品名

                # 金額が有効な数値かチェック
                if not isinstance(amount, (int, float)) or amount <= 0:
                    continue

                # 場所が銀行やローン関連でない場合はスキップ（固定費以外のデータ行を見つける）
                # 実際のデータ行かどうかを判定
                if place and isinstance(place, str):
                    # 日付を生成
                    date = parse_date(sheet_name, day)
                    if not date:
                        continue

                    # カテゴリを自動判定
                    category = categorize_item(place, description)

                    # 商品名・メモ
                    memo = description if description else ''

                    expense = {
                        'date': date,
                        'category': category,
                        'amount': int(amount),
                        'place': place if place else '',
                        'description': str(memo)
                    }

                    expenses.append(expense)

            except Exception as e:
                # エラーは無視して次の行へ
                continue

    wb.close()

    # 日付でソート
    expenses.sort(key=lambda x: x['date'])

    return expenses

def save_to_csv(expenses, output_file):
    """
    CSV形式で保存
    """
    print(f"\nCSVファイルに保存しています: {output_file}")

    with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)

        # ヘッダー
        writer.writerow(['日付', 'カテゴリ', '金額', '場所', '商品名・メモ'])

        # データ
        for expense in expenses:
            writer.writerow([
                expense['date'],
                expense['category'],
                expense['amount'],
                expense['place'],
                expense['description']
            ])

    print(f"✅ {len(expenses)}件のデータを保存しました")

def main():
    input_file = "/Users/runa/Downloads/コピー家計簿　20.06〜 .xlsx"
    output_file = "/Users/runa/kakeibo-app/imported_data.csv"

    print("=" * 60)
    print("家計簿データ変換ツール")
    print("=" * 60)

    try:
        # データ抽出
        expenses = extract_expenses(input_file)

        if not expenses:
            print("\n⚠️ データが見つかりませんでした")
            return

        # CSV保存
        save_to_csv(expenses, output_file)

        # 統計表示
        print("\n" + "=" * 60)
        print("📊 データ統計")
        print("=" * 60)

        # カテゴリ別集計
        category_counts = {}
        for exp in expenses:
            cat = exp['category']
            category_counts[cat] = category_counts.get(cat, 0) + 1

        print("\nカテゴリ別件数:")
        for cat, count in sorted(category_counts.items()):
            print(f"  {cat}: {count}件")

        # 期間
        if expenses:
            first_date = expenses[0]['date']
            last_date = expenses[-1]['date']
            print(f"\n期間: {first_date} 〜 {last_date}")

        print("\n✅ 変換完了！")
        print(f"\n次のステップ:")
        print(f"1. ブラウザでアプリを開く")
        print(f"2. 右上の「データインポート」をクリック")
        print(f"3. {output_file} を選択")
        print(f"4. データがアプリに取り込まれます")

    except Exception as e:
        print(f"\n❌ エラーが発生しました: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()
