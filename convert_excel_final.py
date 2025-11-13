#!/usr/bin/env python3
"""
Excelの家計簿データを年ごとに分割してCSV形式に変換するスクリプト（改良版）
ヘッダー行から列を動的に検出
"""

import openpyxl
import csv
import re
from datetime import datetime
from collections import defaultdict

def categorize_item(place, description):
    """
    場所と商品名からカテゴリを自動判定
    """
    text = f"{place or ''} {description or ''}".lower()

    if any(word in text for word in ['ローン', '保険', 'loan', 'insurance', 'メットライフ',
                                      'アクサ', '車両保険', 'aig', 'ネオファースト']):
        return '固定費'

    if any(word in text for word in ['ソネット', 'モバイル', '携帯', 'スマホ', 'wifi',
                                      'インターネット', 'ネット', '通信', 'ドコモ', 'netflix']):
        return '固定費'

    if any(word in text for word in ['電気', '水道', 'ガス', '光熱', 'でんき', 'みず',
                                      '東京電力', '東京ガス', '下水', '水道局']):
        return '光熱費'

    if any(word in text for word in ['ジム', 'アッティーボ', '映画', 'レジャー', '娯楽']):
        return '娯楽費'

    if any(word in text for word in ['ガソリン', '電車', 'バス', '交通', 'えき', 'タクシー',
                                      'suica', 'パスモ', 'etc']):
        return '交通費'

    if any(word in text for word in ['病院', '薬', '医療', 'クリニック', 'びょういん',
                                      'ドラッグ', '薬局']):
        return '医療費'

    return '食費'

def parse_date(sheet_name, day):
    """
    シート名（YY.MM形式）と日付から完全な日付を生成
    """
    try:
        match = re.match(r'(\d+)\.(\d+)', sheet_name)
        if not match:
            return None

        year_short = int(match.group(1))
        month = int(match.group(2))
        year = 2000 + year_short

        if not isinstance(day, (int, float)):
            return None

        day = int(day)

        if day < 1 or day > 31:
            return None

        date_str = f"{year:04d}-{month:02d}-{day:02d}"

        try:
            datetime.strptime(date_str, '%Y-%m-%d')
            return date_str
        except ValueError:
            return None

    except Exception:
        return None

def find_columns(header_row):
    """
    ヘッダー行から必要な列のインデックスを見つける
    """
    columns = {}

    for i, cell_value in enumerate(header_row):
        if not cell_value:
            continue

        cell_str = str(cell_value).strip()

        if '日' == cell_str or 'day' in cell_str.lower():
            if 'day' not in columns:  # 最初の「日」列のみ
                columns['day'] = i
        elif '場所' in cell_str or 'place' in cell_str.lower():
            columns['place'] = i
        elif '価格' in cell_str or '金額' in cell_str or 'price' in cell_str.lower():
            columns['amount'] = i
        elif '商品' in cell_str or 'item' in cell_str.lower():
            columns['description'] = i

    return columns

def extract_expenses(file_path):
    """
    Excelファイルから支出データを抽出し、年ごとに分類
    """
    print("Excelファイルを読み込んでいます...")
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

    expenses_by_year = defaultdict(list)
    month_sheets = [name for name in wb.sheetnames if re.match(r'\d+\.\d+', name)]

    print(f"\n{len(month_sheets)}個の月次シートを処理します...")

    for sheet_name in month_sheets:
        ws = wb[sheet_name]

        # ヘッダー行から列を検出
        header = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        columns = find_columns(header)

        if not all(k in columns for k in ['day', 'place', 'amount', 'description']):
            print(f"  ⚠️ {sheet_name}: 必要な列が見つかりません - スキップ")
            continue

        print(f"処理中: {sheet_name} (日={columns['day']}, 場所={columns['place']}, 金額={columns['amount']}, 商品={columns['description']})")

        count = 0
        # データ行を処理（最大1000行まで）
        for row in ws.iter_rows(min_row=2, max_row=1000, values_only=True):
            if not row:
                continue

            try:
                # 必要なデータを抽出
                day = row[columns['day']] if len(row) > columns['day'] else None
                place = row[columns['place']] if len(row) > columns['place'] else None
                amount = row[columns['amount']] if len(row) > columns['amount'] else None
                description = row[columns['description']] if len(row) > columns['description'] else None

                # 金額チェック
                if not isinstance(amount, (int, float)) or amount <= 0:
                    continue

                # 日付チェック
                if not isinstance(day, (int, float)):
                    continue

                # 日付を生成
                date = parse_date(sheet_name, day)
                if not date:
                    continue

                # 年を取得
                year = int(date.split('-')[0])

                # データ準備
                place_str = str(place) if place else ''
                desc_str = str(description) if description else ''

                category = categorize_item(place_str, desc_str)

                expense = {
                    'date': date,
                    'category': category,
                    'amount': int(amount),
                    'place': place_str,
                    'description': desc_str
                }

                expenses_by_year[year].append(expense)
                count += 1

            except Exception as e:
                continue

        print(f"  → {count}件のデータを抽出")

    wb.close()

    # 各年のデータを日付でソート
    for year in expenses_by_year:
        expenses_by_year[year].sort(key=lambda x: x['date'])

    return expenses_by_year

def save_to_csv(expenses, output_file):
    """
    CSV形式で保存
    """
    with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(['日付', 'カテゴリ', '金額', '場所', '商品名・メモ'])

        for expense in expenses:
            writer.writerow([
                expense['date'],
                expense['category'],
                expense['amount'],
                expense['place'],
                expense['description']
            ])

def main():
    input_file = "/Users/runa/Downloads/コピー家計簿　20.06〜 .xlsx"
    output_dir = "/Users/runa/kakeibo-app"

    print("=" * 70)
    print("家計簿データ変換ツール (最終版)")
    print("=" * 70)

    try:
        expenses_by_year = extract_expenses(input_file)

        if not expenses_by_year:
            print("\n⚠️ データが見つかりませんでした")
            return

        print("\n" + "=" * 70)
        print("📊 年別データ統計")
        print("=" * 70)

        total_count = 0
        for year in sorted(expenses_by_year.keys()):
            expenses = expenses_by_year[year]
            output_file = f"{output_dir}/imported_data_{year}.csv"

            save_to_csv(expenses, output_file)

            print(f"\n{year}年: {len(expenses)}件")
            print(f"  ファイル: imported_data_{year}.csv")

            total_count += len(expenses)

        print("\n" + "=" * 70)
        print(f"✅ 変換完了！合計 {total_count}件のデータ")
        print("=" * 70)

        print("\n次のステップ:")
        print("1. ブラウザでアプリを開く")
        print("2. 右上の「データインポート」をクリック")
        print("3. インポートしたい年のファイルを選択")
        print("   例: imported_data_2024.csv")
        print("4. 必要に応じて他の年も同様にインポート")

    except Exception as e:
        print(f"\n❌ エラーが発生しました: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()
