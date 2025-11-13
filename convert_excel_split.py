#!/usr/bin/env python3
"""
Excelの家計簿データを年ごとに分割してCSV形式に変換するスクリプト
"""

import openpyxl
import csv
import re
from datetime import datetime
from collections import defaultdict

def categorize_item(place, description):
    """
    場所と商品名からカテゴリを自動判定
    """
    text = f"{place or ''} {description or ''}".lower()

    # ローンや保険などの固定費
    if any(word in text for word in ['ローン', '保険', 'loan', 'insurance', 'メットライフ',
                                      'アクサ', '車両保険', 'aig']):
        return '固定費'

    # 通信費（固定費に含める）
    if any(word in text for word in ['ソネット', 'モバイル', '携帯', 'スマホ', 'wifi',
                                      'インターネット', 'ネット', '通信']):
        return '固定費'

    # 光熱費
    if any(word in text for word in ['電気', '水道', 'ガス', '光熱', 'でんき', 'みず',
                                      '東京電力', '東京ガス', '下水', '水道局']):
        return '光熱費'

    # ジム・娯楽
    if any(word in text for word in ['ジム', 'アッティーボ', '映画', 'レジャー', '娯楽']):
        return '娯楽費'

    # 交通費
    if any(word in text for word in ['ガソリン', '電車', 'バス', '交通', 'えき', 'タクシー',
                                      'suica', 'パスモ']):
        return '交通費'

    # 医療費
    if any(word in text for word in ['病院', '薬', '医療', 'クリニック', 'びょういん',
                                      'ドラッグ', '薬局']):
        return '医療費'

    # 食費（デフォルト）
    return '食費'

def parse_date(sheet_name, day, year_offset=0):
    """
    シート名（YY.MM形式）と日付から完全な日付を生成
    """
    try:
        match = re.match(r'(\d+)\.(\d+)', sheet_name)
        if not match:
            return None

        year_short = int(match.group(1))
        month = int(match.group(2))
        year = 2000 + year_short + year_offset

        if not isinstance(day, (int, float)):
            return None

        day = int(day)

        if day < 1 or day > 31:
            return None

        date_str = f"{year:04d}-{month:02d}-{day:02d}"

        try:
            datetime.strptime(date_str, '%Y-%m-%d')
            return date_str
        except ValueError:
            return None

    except Exception as e:
        return None

def extract_expenses(file_path):
    """
    Excelファイルから支出データを抽出し、年ごとに分類
    """
    print("Excelファイルを読み込んでいます...")
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

    expenses_by_year = defaultdict(list)

    month_sheets = [name for name in wb.sheetnames
                   if re.match(r'\d+\.\d+', name)]

    print(f"\n{len(month_sheets)}個の月次シートを処理します...")

    for sheet_name in month_sheets:
        print(f"処理中: {sheet_name}")
        ws = wb[sheet_name]

        # 全ての行をスキャン（最大1000行まで）
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=1000, values_only=True), start=2):
            if not row:
                continue

            try:
                # 行の長さをチェック
                if len(row) < 8:
                    continue

                day = row[3]
                place = row[5]
                amount = row[6]
                description = row[7] if len(row) > 7 else None

                # 金額が有効かチェック
                if not isinstance(amount, (int, float)) or amount <= 0:
                    continue

                # 日付が有効かチェック
                if not isinstance(day, (int, float)):
                    continue

                # 場所があるかチェック（空でも可）
                date = parse_date(sheet_name, day)
                if not date:
                    continue

                # 年を取得
                year = int(date.split('-')[0])

                # placeがNoneの場合は空文字に
                place_str = place if place else ''

                category = categorize_item(place_str, description)
                memo = description if description else ''

                expense = {
                    'date': date,
                    'category': category,
                    'amount': int(amount),
                    'place': str(place_str),
                    'description': str(memo)
                }

                expenses_by_year[year].append(expense)

            except Exception as e:
                # エラーは無視して次の行へ
                continue

    wb.close()

    # 各年のデータを日付でソート
    for year in expenses_by_year:
        expenses_by_year[year].sort(key=lambda x: x['date'])

    return expenses_by_year

def save_to_csv(expenses, output_file):
    """
    CSV形式で保存
    """
    with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(['日付', 'カテゴリ', '金額', '場所', '商品名・メモ'])

        for expense in expenses:
            writer.writerow([
                expense['date'],
                expense['category'],
                expense['amount'],
                expense['place'],
                expense['description']
            ])

def main():
    input_file = "/Users/runa/Downloads/コピー家計簿　20.06〜 .xlsx"
    output_dir = "/Users/runa/kakeibo-app"

    print("=" * 60)
    print("家計簿データ変換ツール (年別分割版)")
    print("=" * 60)

    try:
        # データ抽出
        expenses_by_year = extract_expenses(input_file)

        if not expenses_by_year:
            print("\n⚠️ データが見つかりませんでした")
            return

        print("\n" + "=" * 60)
        print("📊 年別データ統計")
        print("=" * 60)

        # 各年ごとにCSV保存
        for year in sorted(expenses_by_year.keys()):
            expenses = expenses_by_year[year]
            output_file = f"{output_dir}/imported_data_{year}.csv"

            save_to_csv(expenses, output_file)

            print(f"\n{year}年: {len(expenses)}件")
            print(f"  ファイル: imported_data_{year}.csv")

        print("\n" + "=" * 60)
        print("✅ 変換完了！")
        print("=" * 60)

        print("\n次のステップ:")
        print("1. ブラウザでアプリを開く")
        print("2. 右上の「データインポート」をクリック")
        print("3. インポートしたい年のファイルを選択")
        print("   例: imported_data_2024.csv")
        print("4. 必要に応じて他の年も同様にインポート")
        print("\n💡 ヒント: 最新の年から順にインポートすることをおすすめします")

    except Exception as e:
        print(f"\n❌ エラーが発生しました: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()
