#!/usr/bin/env python3
"""
Excelの家計簿データをCSV形式に変換するスクリプト（大項目・小項目対応）
年ごとに分割して出力
"""

import openpyxl
import csv
import re
from datetime import datetime
from collections import defaultdict

def find_columns(header_row):
    """
    ヘッダー行から列のインデックスを見つける
    """
    columns = {
        'categories': [],  # カテゴリ列（大項目）のリスト
        'subcategories': {}  # サブカテゴリ列（小項目）のマッピング
    }

    # 大項目カテゴリ名のリスト
    main_categories = [
        '食品', '日用品', '外食費', '衣類', '家具・家電', '美容',
        '医療費', '交際費', 'レジャー', 'ガソリン・ETC', '光熱費',
        '通信費', '保険', '車関連【車検・税金・積立】', '税金', '経費', 'ローン'
    ]

    for i, cell_value in enumerate(header_row):
        if not cell_value:
            continue

        cell_str = str(cell_value).strip()

        # 基本列の検出
        if '日' == cell_str or 'day' in cell_str.lower():
            if 'day' not in columns:
                columns['day'] = i
        elif '場所' in cell_str or 'place' in cell_str.lower():
            columns['place'] = i
        elif '価格' in cell_str or '金額' in cell_str or 'price' in cell_str.lower():
            columns['amount'] = i
        elif '商品' in cell_str or 'item' in cell_str.lower():
            columns['description'] = i
        # 大項目カテゴリ列の検出
        elif cell_str in main_categories:
            columns['categories'].append({'index': i, 'name': cell_str})
        # 小項目カテゴリ列の検出（「内訳」を含む列）
        elif '内訳' in cell_str:
            # 「食品内訳」→「食品」のようにマッピング
            main_cat = cell_str.replace('内訳', '')
            columns['subcategories'][main_cat] = i

    return columns

def extract_category_and_subcategory(row, category_columns, subcategory_columns):
    """
    行から大項目カテゴリと小項目を抽出
    """
    main_category = 'その他'
    subcategory = ''
    cat_column_value = None

    # 大項目カテゴリを検出
    for cat_info in category_columns:
        idx = cat_info['index']
        if len(row) > idx and row[idx] is not None:
            value = row[idx]
            if value and str(value).strip():
                main_category = cat_info['name']
                cat_column_value = str(value).strip()
                break

    # 小項目を検出
    # 方法1: カテゴリ列自体の値（文字列の場合）
    if cat_column_value and not cat_column_value.replace('.', '').replace('-', '').isdigit():
        subcategory = cat_column_value

    # 方法2: 「○○内訳」列の値（こちらを優先）
    if main_category != 'その他' and main_category in subcategory_columns:
        sub_idx = subcategory_columns[main_category]
        if len(row) > sub_idx and row[sub_idx] is not None:
            sub_value = row[sub_idx]
            # 数値でない場合のみ小項目として扱う
            if sub_value and str(sub_value).strip():
                sub_str = str(sub_value).strip()
                # 数値や日付でない場合のみ
                if not sub_str.replace('.', '').replace('-', '').replace(':', '').replace(' ', '').isdigit():
                    subcategory = sub_str

    return main_category, subcategory

def parse_date(sheet_name, day):
    """
    シート名（YY.MM形式）と日付から完全な日付を生成
    """
    try:
        match = re.match(r'(\d+)\.(\d+)', sheet_name)
        if not match:
            return None, None, None

        year_short = int(match.group(1))
        month = int(match.group(2))
        year = 2000 + year_short

        if not isinstance(day, (int, float)):
            return None, None, None

        day = int(day)

        if day < 1 or day > 31:
            return None, None, None

        date_str = f"{year:04d}-{month:02d}-{day:02d}"

        try:
            datetime.strptime(date_str, '%Y-%m-%d')
            return date_str, year, month
        except ValueError:
            return None, None, None

    except Exception:
        return None, None, None

def extract_expenses(file_path):
    """
    Excelファイルから支出データを抽出し、年ごとに分類
    """
    print("Excelファイルを読み込んでいます...")
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

    expenses_by_year = defaultdict(list)
    month_sheets = [name for name in wb.sheetnames if re.match(r'\d+\.\d+', name)]

    print(f"\n{len(month_sheets)}個の月次シートを処理します...")

    all_categories = set()
    all_subcategories = set()

    for sheet_name in month_sheets:
        ws = wb[sheet_name]

        # ヘッダー行から列を検出
        header = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        columns = find_columns(header)

        if not all(k in columns for k in ['day', 'place', 'amount', 'description']):
            print(f"  ⚠️ {sheet_name}: 必要な列が見つかりません - スキップ")
            continue

        category_names = [c['name'] for c in columns['categories']]
        print(f"処理中: {sheet_name}")
        print(f"  大項目: {', '.join(category_names)}")
        print(f"  小項目: {', '.join(columns['subcategories'].keys())}")

        count = 0
        # データ行を処理
        for row in ws.iter_rows(min_row=2, max_row=1000, values_only=True):
            if not row:
                continue

            try:
                # 必要なデータを抽出
                day = row[columns['day']] if len(row) > columns['day'] else None
                place = row[columns['place']] if len(row) > columns['place'] else None
                amount = row[columns['amount']] if len(row) > columns['amount'] else None
                description = row[columns['description']] if len(row) > columns['description'] else None

                # 金額チェック
                if not isinstance(amount, (int, float)) or amount <= 0:
                    continue

                # 日付チェック
                if not isinstance(day, (int, float)):
                    continue

                # 日付を生成
                date, year, month = parse_date(sheet_name, day)
                if not date:
                    continue

                # カテゴリと小項目を抽出
                category, subcategory = extract_category_and_subcategory(
                    row, columns['categories'], columns['subcategories']
                )
                all_categories.add(category)
                if subcategory:
                    all_subcategories.add(subcategory)

                # データ準備
                place_str = str(place) if place else ''
                desc_str = str(description) if description else ''

                expense = {
                    'date': date,
                    'category': category,
                    'subcategory': subcategory,
                    'amount': int(amount),
                    'place': place_str,
                    'description': desc_str
                }

                expenses_by_year[year].append(expense)
                count += 1

            except Exception as e:
                continue

        print(f"  → {count}件のデータを抽出")

    wb.close()

    # 各年のデータを日付でソート
    for year in expenses_by_year:
        expenses_by_year[year].sort(key=lambda x: x['date'])

    print(f"\n検出された大項目: {', '.join(sorted(all_categories))}")
    print(f"検出された小項目の例: {', '.join(list(sorted(all_subcategories))[:20])}")

    return expenses_by_year

def save_to_csv(expenses, output_file):
    """
    CSV形式で保存（大項目・小項目を含む）
    """
    with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(['日付', 'カテゴリ', '小項目', '金額', '場所', '商品名・メモ'])

        for expense in expenses:
            writer.writerow([
                expense['date'],
                expense['category'],
                expense['subcategory'],
                expense['amount'],
                expense['place'],
                expense['description']
            ])

def main():
    input_file = "/Users/runa/Downloads/家計簿　20.06〜.xlsx"
    output_dir = "/Users/runa/kakeibo-app"

    print("=" * 70)
    print("家計簿データ変換ツール（大項目・小項目対応版）")
    print("=" * 70)

    try:
        expenses_by_year = extract_expenses(input_file)

        if not expenses_by_year:
            print("\n⚠️ データが見つかりませんでした")
            return

        print("\n" + "=" * 70)
        print("📊 年別データ統計")
        print("=" * 70)

        total_count = 0
        for year in sorted(expenses_by_year.keys()):
            expenses = sorted(expenses_by_year[year], key=lambda x: x['date'])
            output_file = f"{output_dir}/imported_data_{year}.csv"

            save_to_csv(expenses, output_file)

            print(f"\n{year}年: {len(expenses)}件")
            print(f"  ファイル: imported_data_{year}.csv")

            total_count += len(expenses)

        print("\n" + "=" * 70)
        print(f"✅ 変換完了！合計 {total_count}件のデータ")
        print("=" * 70)

        print("\n次のステップ:")
        print("1. ブラウザでアプリを開く")
        print("2. 右上の「データインポート」をクリック")
        print("3. インポートしたい年のファイルを選択")
        print("4. 月別フィルターで表示を絞り込み")

    except Exception as e:
        print(f"\n❌ エラーが発生しました: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()
