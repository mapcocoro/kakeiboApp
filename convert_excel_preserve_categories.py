#!/usr/bin/env python3
"""
Excelの家計簿データをCSV形式に変換するスクリプト（元のカテゴリを保持）
年月ごとに分割して出力
"""

import openpyxl
import csv
import re
from datetime import datetime
from collections import defaultdict

def find_columns(header_row):
    """
    ヘッダー行から列のインデックスを見つける
    """
    columns = {
        'categories': []  # カテゴリ列のリスト
    }

    # 正しいカテゴリ名のリスト（Excelで使われているもの）
    valid_categories = [
        '食品', '日用品', '外食費', '衣類', '家具・家電', '美容',
        '医療費', '交際費', 'レジャー', 'ガソリン・ETC', '光熱費',
        '通信費', '保険', '車関連【車検・税金・積立】', '税金', '経費', 'ローン'
    ]

    for i, cell_value in enumerate(header_row):
        if not cell_value:
            continue

        cell_str = str(cell_value).strip()

        # 基本列の検出
        if '日' == cell_str or 'day' in cell_str.lower():
            if 'day' not in columns:
                columns['day'] = i
        elif '場所' in cell_str or 'place' in cell_str.lower():
            columns['place'] = i
        elif '価格' in cell_str or '金額' in cell_str or 'price' in cell_str.lower():
            columns['amount'] = i
        elif '商品' in cell_str or 'item' in cell_str.lower():
            columns['description'] = i
        # カテゴリ列の検出（有効なカテゴリ名のみ）
        elif cell_str in valid_categories:
            columns['categories'].append({'index': i, 'name': cell_str})

    return columns

def extract_category(row, category_columns):
    """
    行からカテゴリを抽出
    カテゴリ列のいずれかに値がある場合、そのカテゴリ名を返す
    """
    for cat_info in category_columns:
        idx = cat_info['index']
        if len(row) > idx and row[idx] is not None:
            # 値があればそのカテゴリ（文字列または数値）
            value = row[idx]
            # 空文字列やNoneでなければカテゴリとして認識
            if value and str(value).strip():
                return cat_info['name']

    return 'その他'

def parse_date(sheet_name, day):
    """
    シート名（YY.MM形式）と日付から完全な日付を生成
    """
    try:
        match = re.match(r'(\d+)\.(\d+)', sheet_name)
        if not match:
            return None, None, None

        year_short = int(match.group(1))
        month = int(match.group(2))
        year = 2000 + year_short

        if not isinstance(day, (int, float)):
            return None, None, None

        day = int(day)

        if day < 1 or day > 31:
            return None, None, None

        date_str = f"{year:04d}-{month:02d}-{day:02d}"

        try:
            datetime.strptime(date_str, '%Y-%m-%d')
            return date_str, year, month
        except ValueError:
            return None, None, None

    except Exception:
        return None, None, None

def extract_expenses(file_path):
    """
    Excelファイルから支出データを抽出し、年月ごとに分類
    """
    print("Excelファイルを読み込んでいます...")
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

    expenses_by_year_month = defaultdict(list)
    month_sheets = [name for name in wb.sheetnames if re.match(r'\d+\.\d+', name)]

    print(f"\n{len(month_sheets)}個の月次シートを処理します...")

    all_categories = set()

    for sheet_name in month_sheets:
        ws = wb[sheet_name]

        # ヘッダー行から列を検出
        header = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        columns = find_columns(header)

        if not all(k in columns for k in ['day', 'place', 'amount', 'description']):
            print(f"  ⚠️ {sheet_name}: 必要な列が見つかりません - スキップ")
            continue

        # カテゴリ列の表示
        category_names = [c['name'] for c in columns['categories']]
        print(f"処理中: {sheet_name}")
        print(f"  カテゴリ列: {', '.join(category_names)}")

        count = 0
        # データ行を処理
        for row in ws.iter_rows(min_row=2, max_row=1000, values_only=True):
            if not row:
                continue

            try:
                # 必要なデータを抽出
                day = row[columns['day']] if len(row) > columns['day'] else None
                place = row[columns['place']] if len(row) > columns['place'] else None
                amount = row[columns['amount']] if len(row) > columns['amount'] else None
                description = row[columns['description']] if len(row) > columns['description'] else None

                # 金額チェック
                if not isinstance(amount, (int, float)) or amount <= 0:
                    continue

                # 日付チェック
                if not isinstance(day, (int, float)):
                    continue

                # 日付を生成
                date, year, month = parse_date(sheet_name, day)
                if not date:
                    continue

                # カテゴリを抽出
                category = extract_category(row, columns['categories'])
                all_categories.add(category)

                # データ準備
                place_str = str(place) if place else ''
                desc_str = str(description) if description else ''

                expense = {
                    'date': date,
                    'category': category,
                    'amount': int(amount),
                    'place': place_str,
                    'description': desc_str
                }

                # 年月でグループ化
                year_month_key = f"{year}-{month:02d}"
                expenses_by_year_month[year_month_key].append(expense)
                count += 1

            except Exception as e:
                continue

        print(f"  → {count}件のデータを抽出")

    wb.close()

    # 各月のデータを日付でソート
    for year_month in expenses_by_year_month:
        expenses_by_year_month[year_month].sort(key=lambda x: x['date'])

    print(f"\n検出されたカテゴリ: {', '.join(sorted(all_categories))}")

    return expenses_by_year_month

def save_to_csv(expenses, output_file):
    """
    CSV形式で保存
    """
    with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(['日付', 'カテゴリ', '金額', '場所', '商品名・メモ'])

        for expense in expenses:
            writer.writerow([
                expense['date'],
                expense['category'],
                expense['amount'],
                expense['place'],
                expense['description']
            ])

def main():
    input_file = "/Users/runa/Downloads/コピー家計簿　20.06〜 .xlsx"
    output_dir = "/Users/runa/kakeibo-app"

    print("=" * 70)
    print("家計簿データ変換ツール（カテゴリ保持版）")
    print("=" * 70)

    try:
        expenses_by_year_month = extract_expenses(input_file)

        if not expenses_by_year_month:
            print("\n⚠️ データが見つかりませんでした")
            return

        print("\n" + "=" * 70)
        print("📊 年月別データ統計")
        print("=" * 70)

        total_count = 0

        # 年ごとにまとめて出力
        expenses_by_year = defaultdict(list)
        for year_month, expenses in expenses_by_year_month.items():
            year = year_month.split('-')[0]
            expenses_by_year[year].extend(expenses)

        for year in sorted(expenses_by_year.keys()):
            expenses = sorted(expenses_by_year[year], key=lambda x: x['date'])
            output_file = f"{output_dir}/imported_data_{year}.csv"

            save_to_csv(expenses, output_file)

            print(f"\n{year}年: {len(expenses)}件")
            print(f"  ファイル: imported_data_{year}.csv")

            total_count += len(expenses)

        print("\n" + "=" * 70)
        print(f"✅ 変換完了！合計 {total_count}件のデータ")
        print("=" * 70)

        print("\n次のステップ:")
        print("1. ブラウザでアプリを開く")
        print("2. 右上の「データインポート」をクリック")
        print("3. インポートしたい年のファイルを選択")
        print("4. 月別フィルターで表示を絞り込み")

    except Exception as e:
        print(f"\n❌ エラーが発生しました: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()
